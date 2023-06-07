import openpyxl
from operator import itemgetter

employee_file = openpyxl.load_workbook("employees.xlsx")
employee_list = employee_file["Sheet1"]

job_title_column_index = 3
date_of_birth_column_index = 4
employee_list.delete_cols(job_title_column_index, date_of_birth_column_index)

employee_by_experience = {}

for employee_row in range(2, employee_list.max_row + 1):
    employee_name = employee_list.cell(employee_row, 1).value
    employee_experience = employee_list.cell(employee_row, 2).value

    employee_by_experience[employee_name] = employee_experience

employee_by_experience_sorted = sorted(employee_by_experience.items(), key=lambda item:item[1], reverse=True)

for employee_row in range(2, employee_list.max_row + 1):
    employee_name = employee_list.cell(employee_row, 1)
    employee_experience = employee_list.cell(employee_row, 2)

    index = employee_row - 2
    employee_and_experience = employee_by_experience_sorted[index]

    employee_name.value = employee_and_experience[0]
    employee_experience.value = employee_and_experience[1]

employee_file.save("employee_sorted_by_experience.xlsx")
